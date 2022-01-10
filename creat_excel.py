from openpyxl import Workbook
from datetime import datetime
wb = Workbook();
ws = wb.active;

ws.title="test_1";
ws["A1"]="好好学习,天天向111上";
ws['A2'] = datetime.now()
ws.append([1, 2, 3])

wb.save("creat_excel1.xlsx")
wb.close()
